#!/usr/bin/env python3
"""Structural coverage of the four reference workbooks, not cell equality.

`WF-022` makes this a hard gate, and `WF-033` lists it under pilot-ready gate 2.
The four hand-made workbooks in `data/reference-itineraries/` are four real trips
planned in Excel, and their **recurring sheets are the merged app's entire output
surface** — so comparing against them validates the merge itself. Validation
never compares against Streamlit's output, which is why this reads the workbooks
directly.

Coverage is **structural**: does the app carry a counterpart for each recurring
element? Not cell equality — the workbooks are hand-made and inconsistent, and
the inconsistency is measured rather than assumed:

    ตารางเวลา          4 of 4 workbooks
    ค่าใช้จ่าย           4 of 4
    ☺ Things to Bring  4 of 4
    ♢ To-Do List       3 of 4   (Kunming has none)

Japan carries a `Transport` sheet and Shanghai a `Disney` sheet. Both are
one-offs, so neither is a recurring type and neither is in scope.

## The Thai labels below are citations, and citations get validated

Every marker this gate claims to find in a reference workbook is checked against
the files first. Get a label wrong and the gate fails on the **reference** side,
naming the marker — it does not quietly pass having compared nothing. That is the
same rule the element-parity gate needed: a gate that requires a thing without
validating it is worth nothing, and eight of eleven `derives-from:` citations
were wrong when that check was added.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from io import BytesIO
# A verification script builds throwaway trips, so it must never be pointed at a
# hosted database. See the same guard in `scripts/check.py` and `tests/__init__.py`.
import os as _os

_os.environ.pop("TOURIST_DB_URL", None)

import json
from pathlib import Path
import sys
from tempfile import TemporaryDirectory


ROOT = Path(__file__).resolve().parents[1]
REFERENCES = ROOT / "data" / "reference-itineraries"

TIMETABLE = "ตารางเวลา"
EXPENSES = "ค่าใช้จ่าย"
TODO = "♢ To-Do List"
BRING = "☺ Things to Bring"
RECURRING = (TIMETABLE, EXPENSES, TODO, BRING)

# Sheets that appear in one trip only. Named so their absence from the recurring
# set is a recorded decision rather than an oversight.
ONE_OFF = {"Transport": "japan", "Disney": "shanghai"}


@dataclass(frozen=True)
class Element:
    """One recurring thing a reference sheet carries, and its app counterpart.

    `marker` is text that must appear somewhere in the reference sheet — the
    citation. `app_sheet` and `app_markers` are what must appear in the workbook
    the app generates. `ics` covers the one element the calendar answers for.
    """

    sheet: str
    meaning: str
    marker: str
    app_sheet: str = ""
    app_markers: tuple[str, ...] = ()
    ics: bool = False


ELEMENTS = (
    # --- ตารางเวลา: the day-by-day timetable -----------------------------
    Element(
        sheet=TIMETABLE,
        meaning="a clock time against each activity",
        marker="Time",
        app_sheet="Timeline",
        app_markers=("Start", "End"),
    ),
    Element(
        sheet=TIMETABLE,
        meaning="activities grouped into numbered days",
        marker="Day 1",
        app_sheet="Timeline",
        app_markers=("Date", "Order"),
    ),
    Element(
        sheet=TIMETABLE,
        meaning="what the activity actually is",
        marker="ออกจากบ้านไปสนามบิน",  # "leave home for the airport"
        app_sheet="Timeline",
        app_markers=("Name", "Type"),
    ),
    # --- ค่าใช้จ่าย: the money sheet -------------------------------------
    Element(
        sheet=EXPENSES,
        meaning="a named cost line",
        marker="หัวข้อ",  # "item"
        app_sheet="Costs",
        app_markers=("Cost", "Category"),
    ),
    Element(
        sheet=EXPENSES,
        meaning="what that line cost",
        marker="ค่าใช้จ่าย",
        app_sheet="Costs",
        app_markers=("Original amount", "Original currency"),
    ),
    Element(
        sheet=EXPENSES,
        meaning="the trip total",
        marker="ค่าใช้จ่ายรวม",
        app_sheet="Costs",
        app_markers=("Total THB",),
    ),
    Element(
        sheet=EXPENSES,
        meaning="the cost per person",
        marker="ค่าใช้จ่ายต่อคน",
        app_sheet="Costs",
        app_markers=("Per person THB",),
    ),
    Element(
        sheet=EXPENSES,
        meaning="whether a line is already paid",
        marker="จ่ายแล้ว",
        app_sheet="Costs",
        app_markers=("Payment state", "Paid THB"),
    ),
    # --- ♢ To-Do List: things to arrange before leaving -------------------
    Element(
        sheet=TODO,
        meaning="a pre-trip task list with progress",
        marker="IMPORTANT TO-DO LIST",
        app_sheet="Checklist",
        app_markers=("Title", "Progress", "Due date / milestone"),
        ics=True,
    ),
    # --- ☺ Things to Bring: what to pack ---------------------------------
    Element(
        sheet=BRING,
        meaning="a packing list that can be ticked off",
        marker="MY CHECKLIST",
        app_sheet="Checklist",
        app_markers=("Category", "packing"),
    ),
)


@dataclass
class Result:
    inventory: dict[str, list[str]] = field(default_factory=dict)
    uncited: list[str] = field(default_factory=list)
    uncovered: list[str] = field(default_factory=list)
    covered: int = 0


def reference_text() -> tuple[dict[str, list[str]], dict[str, str]]:
    """Which workbooks hold which sheet, and all text per sheet name."""

    from openpyxl import load_workbook

    inventory: dict[str, list[str]] = {name: [] for name in RECURRING}
    text: dict[str, list[str]] = {name: [] for name in RECURRING}
    for path in sorted(REFERENCES.glob("*/*.xlsx")):
        trip = path.parent.name
        workbook = load_workbook(path, data_only=True)
        try:
            for name in workbook.sheetnames:
                if name not in RECURRING:
                    continue
                inventory[name].append(trip)
                sheet = workbook[name]
                for row in sheet.iter_rows(values_only=True):
                    text[name].extend(str(cell) for cell in row if cell is not None)
        finally:
            workbook.close()
    return inventory, {name: "\n".join(values) for name, values in text.items()}


def app_workbook() -> tuple[dict[str, str], str]:
    """Generate one populated workbook and calendar, and return their text.

    Driven through the historic fixture and the real actions layer, so this is
    the same output a download would serve. No network, no paid call, and no
    dependence on `data/tourist.sqlite3`.
    """

    from openpyxl import load_workbook
    from unittest.mock import patch

    sys.path.insert(0, str(ROOT))
    from travel_planner.actions import PlannerActions
    from travel_planner.core import new_optimization_preview
    from travel_planner.exporters import checklist_ics, plan_workbook_xlsx
    from travel_planner.optimizer import optimize_trip

    catalog = json.loads(
        (ROOT / "tests" / "fixtures" / "historic_regressions.json").read_text(
            encoding="utf-8"
        )
    )
    fixture = next(
        item
        for item in catalog["fixtures"]
        if item["metadata"]["id"] == "ix-jp-shibuya-hours-view-walk"
    )
    planner_input = json.loads(json.dumps(fixture["planner_input"]))
    proposal = optimize_trip(planner_input)

    with TemporaryDirectory() as directory:
        actions = PlannerActions(Path(directory) / "coverage.sqlite3")
        trip = actions.create_trip(name="Coverage probe", destination="Tokyo")
        # A confirmed setup is what lets the readiness board generate; the board
        # itself is city-independent, so no discovery or provider is needed.
        # The dates matter: a readiness item's deadline is relative to the trip,
        # and `checklist_ics` only emits an event for a **dated** item. Without
        # them the calendar is legitimately empty and this gate would report an
        # app gap that is really a gap in its own probe.
        day = planner_input["trip"]["local_dates"][0]
        actions.save_setup(
            trip_id=trip.trip_id,
            main_style=["sightseeing"],
            travellers=[{"label": "Companion", "age": 30, "tags": []}],
            start_date=day,
            end_date=day,
            accommodation_status="booked",
            confirmed=True,
        )
        actions.store.save_optimization_preview(
            new_optimization_preview(
                trip_id=trip.trip_id, optimizer_input=planner_input, proposal=proposal
            )
        )
        with patch.object(actions, "_optimizer_input", return_value=planner_input):
            actions.activate_plan_preview(
                trip_id=trip.trip_id, variant_id="best_balance"
            )
        actions.apply_checklist_proposal(trip.trip_id)
        actions.save_rate_snapshot(
            trip_id=trip.trip_id,
            rates={"JPY": 0.24},
            as_of="2026-08-01",
            source="coverage_probe",
        )
        # One paid line and one estimate, so the totals block has both to report.
        actions.save_cost_item(
            trip_id=trip.trip_id,
            item={
                "label": "Flights",
                "category": "transport",
                "original_amount": 15429.52,
                "original_currency": "THB",
                "payment_state": "paid",
                "actual_thb": 15429.52,
            },
        )
        actions.save_cost_item(
            trip_id=trip.trip_id,
            item={
                "label": "Hotel",
                "category": "accommodation",
                "original_amount": 40000,
                "original_currency": "JPY",
                "payment_state": "estimate",
            },
        )
        snapshot = actions.build_export_snapshot(trip.trip_id).as_dict()

    workbook = load_workbook(BytesIO(plan_workbook_xlsx(snapshot)), data_only=True)
    try:
        text = {}
        for name in workbook.sheetnames:
            sheet = workbook[name]
            values = []
            for row in sheet.iter_rows(values_only=True):
                values.extend(str(cell) for cell in row if cell is not None)
            text[name] = "\n".join(values)
    finally:
        workbook.close()
    return text, checklist_ics(snapshot).decode("utf-8")


def main() -> int:
    if not REFERENCES.is_dir() or not any(REFERENCES.glob("*/*.xlsx")):
        print("SKIP: no reference workbooks in this checkout", flush=True)
        return 0
    try:
        import openpyxl  # noqa: F401
    except ModuleNotFoundError:
        print(
            "FAILED: openpyxl is needed to read the reference workbooks; it is in "
            "the dev group",
            file=sys.stderr,
        )
        return 1

    inventory, reference = reference_text()
    sheets, calendar = app_workbook()
    result = Result(inventory=inventory)

    for element in ELEMENTS:
        # The citation first: if the marker is not in the references, this gate
        # is describing a workbook that does not exist.
        if element.marker not in reference.get(element.sheet, ""):
            result.uncited.append(
                f"{element.sheet}: {element.marker!r} is cited for "
                f"{element.meaning!r} but appears in no reference workbook"
            )
            continue
        missing = [
            marker
            for marker in element.app_markers
            if marker not in sheets.get(element.app_sheet, "")
        ]
        if element.ics and "BEGIN:VEVENT" not in calendar:
            missing.append("a calendar event")
        if missing:
            result.uncovered.append(
                f"{element.sheet} → {element.app_sheet}: {element.meaning} — "
                f"nothing covers {', '.join(repr(m) for m in missing)}"
            )
        else:
            result.covered += 1

    print(f"  reference workbooks read: {len(list(REFERENCES.glob('*/*.xlsx')))}", flush=True)
    for name in RECURRING:
        trips = inventory.get(name) or []
        note = "" if len(trips) == 4 else f"  (absent from {4 - len(trips)})"
        print(f"    {name}: {len(trips)} of 4{note}", flush=True)
    print(f"  one-off sheets, not recurring, out of scope: "
          f"{', '.join(f'{k} ({v})' for k, v in ONE_OFF.items())}", flush=True)
    print(f"  recurring elements covered: {result.covered} of {len(ELEMENTS)}", flush=True)

    if result.uncited:
        print(
            f"FAILED: {len(result.uncited)} citation(s) do not match the reference "
            "workbooks — fix the gate, not the app",
            file=sys.stderr,
        )
        for problem in result.uncited:
            print(f"  - {problem}", file=sys.stderr)
        return 1
    if result.uncovered:
        print(
            f"FAILED: {len(result.uncovered)} recurring element(s) have no counterpart "
            "in the app's output",
            file=sys.stderr,
        )
        for problem in result.uncovered:
            print(f"  - {problem}", file=sys.stderr)
        return 1
    print(
        "PASS: every recurring element of the four reference workbooks has a "
        "structural counterpart",
        flush=True,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
